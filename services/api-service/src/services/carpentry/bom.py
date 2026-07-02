from datetime import date, datetime
from io import BytesIO
import re
import unicodedata

from sqlalchemy.ext.asyncio import AsyncSession

from src.core.logging_config import get_logger
from src.services.carpentry.common import AppError, clean, decimal_or_none, execute, fetch_all, fetch_one, to_bool

logger = get_logger(__name__)


def _unidad_por_categoria(categoria: str | None) -> str:
    if categoria == "tablero":
        return "m2"
    if categoria == "canto":
        return "m"
    if categoria == "herraje":
        return "unidad"
    return "unidad"


def _norm(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _text_or_none(value) -> str | None:
    value = clean(value)
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_key(value) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("milimetros", "mm").replace("milimetro", "mm")
    text = re.sub(r"\b(\d+)\s*mm\b", r"\1", text)
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def _split_description(value: str | None) -> list[str]:
    return [part.strip() for part in re.split(r"[-|]", str(value or "")) if part and part.strip()]


async def ensure_items_proyecto_schema(db: AsyncSession) -> None:
    await execute(
        db,
        """ALTER TABLE items_proyecto
           ADD COLUMN IF NOT EXISTS tipologia TEXT""",
    )
    await execute(
        db,
        """ALTER TABLE items_proyecto
           ADD COLUMN IF NOT EXISTS apartamento TEXT""",
    )
    await execute(
        db,
        """DO $$
           BEGIN
             IF EXISTS (
               SELECT 1
               FROM information_schema.columns
               WHERE table_name = 'items_proyecto'
                 AND column_name = 'piso'
             ) THEN
               UPDATE items_proyecto
               SET tipologia = COALESCE(tipologia, piso::text)
               WHERE tipologia IS NULL
                 AND piso IS NOT NULL;
             END IF;
           END $$""",
    )


def _parse_material_from_quote(tipo: str | None, descripcion: str | None) -> dict:
    categoria = _normalize_material_category(tipo)
    parts = _split_description(descripcion)
    raw = str(descripcion or "").strip()

    if categoria == "tablero":
        calibre = None
        material = parts[0] if parts else raw
        sustrato = parts[1] if len(parts) > 1 else None
        if parts:
            last = parts[-1]
            match = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:mm)?$", last, re.IGNORECASE)
            if match:
                calibre = f"{match.group(1).replace(',', '.')} mm"
        return {
            "categoria": "tablero",
            "nombre": raw,
            "material": material,
            "sustrato": sustrato,
            "calibre": calibre,
            "unidad_medida": "m2",
        }

    if categoria == "canto":
        sustrato = parts[0] if parts else raw
        proveedor = parts[1] if len(parts) > 1 else None
        return {
            "categoria": "canto",
            "nombre": raw,
            "sustrato": sustrato,
            "proveedor": proveedor,
            "unidad_medida": "m",
        }

    return {
        "categoria": "herraje",
        "nombre": raw,
        "referencia": raw,
        "unidad_medida": "unidad",
    }


def _normalize_material_category(value: str | None) -> str:
    normalized = _normalize_key(value)
    if "tablero" in normalized:
        return "tablero"
    if "canto" in normalized:
        return "canto"
    return "herraje"


def _calibre_con_mm(value):
    calibre = clean(value)
    if not calibre:
        return None
    without_mm = str(calibre).replace("mm", "").replace("MM", "").strip()
    return f"{without_mm} mm" if without_mm else None


def _int_or_none(value, field_name: str) -> int | None:
    value = clean(value)
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise AppError(f"Valor inválido para {field_name}.", 400, "VALIDATION") from exc


def _int_or_zero(value) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _date_or_none(value, field_name: str) -> date | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text[:10]).date()
    except ValueError as exc:
        raise AppError(f"Fecha inválida en {field_name}.", 400, "VALIDATION") from exc


def _nombre_canonico_por_categoria(data: dict) -> str | None:
    categoria = data.get("categoria")

    if categoria == "tablero":
        material = clean(data.get("material")) or clean(data.get("nombre"))
        calibre = _calibre_con_mm(data.get("calibre"))
        if not material:
            return None
        return f"{material} - {calibre}" if calibre else material

    if categoria == "canto":
        sustrato = clean(data.get("sustrato")) or clean(data.get("nombre"))
        calibre = _calibre_con_mm(data.get("calibre"))
        proveedor = clean(data.get("proveedor"))
        if not sustrato:
            return None
        base = f"{sustrato} - {calibre}" if calibre else sustrato
        return f"{base} | {proveedor}" if proveedor else base

    if categoria == "herraje":
        return clean(data.get("referencia")) or clean(data.get("nombre"))

    return clean(data.get("nombre"))


async def _find_material_id_by_variant(db: AsyncSession, data: dict) -> int | None:
    categoria = data.get("categoria")

    if categoria == "tablero":
        row = await fetch_one(
            db,
            """SELECT mc.id
               FROM materiales_catalogo mc
               JOIN materiales_tableros mt ON mt.material_id = mc.id
               WHERE mc.categoria = 'tablero'
                 AND LOWER(COALESCE(mt.material, '')) = LOWER(COALESCE($1, ''))
                 AND LOWER(COALESCE(mt.sustrato, '')) = LOWER(COALESCE($2, ''))
                 AND LOWER(COALESCE(mt.formato, '')) = LOWER(COALESCE($3, ''))
                 AND LOWER(COALESCE(mt.calibre, '')) = LOWER(COALESCE($4, ''))
                 AND LOWER(COALESCE(mt.proveedor, '')) = LOWER(COALESCE($5, ''))
               ORDER BY mc.id ASC
               LIMIT 1""",
            [
                _norm(data.get("material")) or _norm(data.get("nombre")),
                _norm(data.get("sustrato")),
                _norm(data.get("formato")),
                _norm(data.get("calibre")),
                _norm(data.get("proveedor")),
            ],
        )
        return row["id"] if row else None

    if categoria == "canto":
        row = await fetch_one(
            db,
            """SELECT mc.id
               FROM materiales_catalogo mc
               JOIN materiales_cantos mca ON mca.material_id = mc.id
               WHERE mc.categoria = 'canto'
                 AND LOWER(COALESCE(mca.sustrato, '')) = LOWER(COALESCE($1, ''))
                 AND LOWER(COALESCE(mca.calibre, '')) = LOWER(COALESCE($2, ''))
                 AND LOWER(COALESCE(mca.proveedor, '')) = LOWER(COALESCE($3, ''))
               ORDER BY mc.id ASC
               LIMIT 1""",
            [
                _norm(data.get("sustrato")) or _norm(data.get("nombre")),
                _norm(data.get("calibre")),
                _norm(data.get("proveedor")),
            ],
        )
        return row["id"] if row else None

    if categoria == "herraje":
        row = await fetch_one(
            db,
            """SELECT mc.id
               FROM materiales_catalogo mc
               JOIN materiales_herrajes mh ON mh.material_id = mc.id
               WHERE mc.categoria = 'herraje'
                 AND LOWER(COALESCE(mh.referencia, '')) = LOWER(COALESCE($1, ''))
                 AND LOWER(COALESCE(mh.proveedor, '')) = LOWER(COALESCE($2, ''))
               ORDER BY mc.id ASC
               LIMIT 1""",
            [
                _norm(data.get("referencia")) or _norm(data.get("nombre")),
                _norm(data.get("proveedor")),
            ],
        )
        return row["id"] if row else None

    return None


async def listar_items_por_proyecto(db: AsyncSession, payload: dict | None = None) -> list[dict]:
    payload = payload or {}
    proyecto_id = _int_or_none(payload.get("proyecto_id"), "proyecto_id")

    if not proyecto_id:
        return []

    await ensure_items_proyecto_schema(db)

    return await fetch_all(
        db,
        """SELECT ip.id, ip.proyecto_id, ip.lote_id,
                  l.nombre AS lote,
                  ip.nombre, ip.tipologia, ip.tipologia AS piso, ip.apartamento, ip.estado
           FROM items_proyecto ip
           LEFT JOIN lotes l ON l.id = ip.lote_id
           WHERE ip.proyecto_id = $1
           ORDER BY ip.nombre ASC, ip.tipologia ASC NULLS LAST, ip.id ASC""",
        [proyecto_id],
    )


async def crear_item(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    await ensure_items_proyecto_schema(db)
    data = {
        "proyecto_id": _int_or_none(payload.get("proyecto_id"), "proyecto_id"),
        "nombre": clean(payload.get("nombre")),
        "tipologia": _text_or_none(payload.get("tipologia")) or _text_or_none(payload.get("piso")),
        "apartamento": clean(payload.get("apartamento")),
    }

    if not data["proyecto_id"] or not data["nombre"]:
        raise AppError("Proyecto y nombre del ítem son obligatorios.", 400, "VALIDATION")

    async with db.begin():
        rows = await execute(
            db,
            """INSERT INTO items_proyecto (proyecto_id, nombre, tipologia, apartamento)
               VALUES ($1, $2, $3, $4)
               RETURNING id""",
            [data["proyecto_id"], data["nombre"], data["tipologia"], data["apartamento"]],
        )

    return {"id": rows[0]["id"], "creado": True}


async def actualizar_item(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    await ensure_items_proyecto_schema(db)
    data = {
        "id": _int_or_none(payload.get("id"), "id"),
        "nombre": clean(payload.get("nombre")),
        "tipologia": _text_or_none(payload.get("tipologia")) or _text_or_none(payload.get("piso")),
        "apartamento": clean(payload.get("apartamento")),
    }

    if not data["id"] or not data["nombre"]:
        raise AppError("ID y nombre del ítem son obligatorios.", 400, "VALIDATION")

    async with db.begin():
        rows = await execute(
            db,
            """UPDATE items_proyecto
               SET nombre = $1,
                   tipologia = $2,
                   apartamento = $3
               WHERE id = $4
               RETURNING id""",
            [data["nombre"], data["tipologia"], data["apartamento"], data["id"]],
        )

    if not rows:
        raise AppError("Ítem no encontrado.", 404, "NOT_FOUND")

    return {"id": rows[0]["id"], "creado": False}


async def eliminar_item(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    item_id = _int_or_none(payload.get("id"), "id")

    if not item_id:
        raise AppError("ID del ítem es obligatorio.", 400, "VALIDATION")

    async with db.begin():
        bom_check = await fetch_one(
            db,
            """SELECT COUNT(*)::INTEGER AS total
               FROM bom_items
               WHERE item_id = $1""",
            [item_id],
        )

        if bom_check and bom_check["total"] > 0:
            raise AppError(
                "No se puede eliminar el ítem porque tiene materiales BOM asociados. Elimina el BOM primero.",
                400,
                "HAS_BOM",
            )

        await execute(db, "DELETE FROM items_proyecto WHERE id = $1", [item_id])

    return {"ok": True}


async def asignar_lote(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    item_id = _int_or_none(payload.get("id"), "id")
    lote_id = _int_or_none(payload.get("lote_id"), "lote_id")

    if not item_id or not lote_id:
        raise AppError("ID del ítem y del lote son obligatorios.", 400, "VALIDATION")

    async with db.begin():
        rows = await execute(
            db,
            """UPDATE items_proyecto
               SET lote_id = $1,
                   estado = 'en_lote'
               WHERE id = $2
               RETURNING id""",
            [lote_id, item_id],
        )

    if not rows:
        raise AppError("Ítem no encontrado.", 404, "NOT_FOUND")

    return {"ok": True}


async def desasignar_lote(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    item_id = _int_or_none(payload.get("id"), "id")

    if not item_id:
        raise AppError("ID del ítem es obligatorio.", 400, "VALIDATION")

    async with db.begin():
        rows = await execute(
            db,
            """UPDATE items_proyecto
               SET lote_id = NULL,
                   estado = 'pendiente'
               WHERE id = $1
               RETURNING id""",
            [item_id],
        )

    if not rows:
        raise AppError("Ítem no encontrado.", 404, "NOT_FOUND")

    return {"ok": True}


async def copiar_bom_a_items(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    item_origen_id = _int_or_none(payload.get("item_origen_id"), "item_origen_id")
    raw_destinos = payload.get("item_destinos_ids")
    item_destinos_ids: list[int] = []
    if isinstance(raw_destinos, list):
        for item in raw_destinos:
            parsed = _int_or_none(item, "item_destinos_ids")
            if parsed is not None:
                item_destinos_ids.append(parsed)

    if not item_origen_id or not isinstance(item_destinos_ids, list) or not item_destinos_ids:
        raise AppError(
            "item_origen_id y al menos un item_destinos_ids son obligatorios.",
            400,
            "VALIDATION",
        )

    async with db.begin():
        bom_origen = await fetch_all(
            db,
            """SELECT material_id, cantidad_requerida, notas
               FROM bom_items
               WHERE item_id = $1""",
            [item_origen_id],
        )

        if not bom_origen:
            raise AppError("El ítem origen no tiene materiales BOM para copiar.", 400, "EMPTY_BOM")

        item_ids = []
        material_ids = []
        cantidades = []
        notas = []

        for destino_id in item_destinos_ids:
            for row in bom_origen:
                item_ids.append(destino_id)
                material_ids.append(row["material_id"])
                cantidades.append(row["cantidad_requerida"])
                notas.append(row["notas"])

        await execute(
            db,
            """INSERT INTO bom_items (item_id, material_id, cantidad_requerida, notas)
               SELECT *
               FROM unnest(
                 $1::INT[],
                 $2::INT[],
                 $3::NUMERIC[],
                 $4::TEXT[]
               ) AS src(item_id, material_id, cantidad_requerida, notas)
               ON CONFLICT (item_id, material_id)
               DO UPDATE SET cantidad_requerida = EXCLUDED.cantidad_requerida,
                             notas = EXCLUDED.notas""",
            [item_ids, material_ids, cantidades, notas],
        )

    return {"ok": True, "copiados": len(item_destinos_ids)}


async def listar_bom_por_item(db: AsyncSession, payload: dict | None = None) -> list[dict]:
    payload = payload or {}
    item_id = _int_or_none(payload.get("item_id"), "item_id")
    if not item_id:
        return []

    return await fetch_all(
        db,
        """SELECT bi.id, bi.item_id, bi.material_id,
                  mc.nombre AS material, mc.categoria, mc.unidad_medida,
                  bi.cantidad_requerida, bi.notas
           FROM bom_items bi
           JOIN materiales_catalogo mc ON mc.id = bi.material_id
           WHERE bi.item_id = $1
           ORDER BY mc.nombre""",
        [item_id],
    )


async def guardar_material_bom(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    data = {
        "item_id": _int_or_none(payload.get("item_id"), "item_id"),
        "material_id": _int_or_none(payload.get("material_id"), "material_id"),
        "cantidad_requerida": decimal_or_none(payload.get("cantidad_requerida")),
        "notas": clean(payload.get("notas")),
    }

    if not data["item_id"] or not data["material_id"] or not data["cantidad_requerida"] or data["cantidad_requerida"] <= 0:
        raise AppError(
            "Item, material y cantidad requerida (>0) son obligatorios.",
            400,
            "VALIDATION",
        )

    async with db.begin():
        rows = await execute(
            db,
            """INSERT INTO bom_items (item_id, material_id, cantidad_requerida, notas)
               VALUES ($1,$2,$3,$4)
               ON CONFLICT (item_id, material_id)
               DO UPDATE SET cantidad_requerida = EXCLUDED.cantidad_requerida,
                             notas = EXCLUDED.notas
               RETURNING id""",
            [
                data["item_id"],
                data["material_id"],
                data["cantidad_requerida"],
                data["notas"],
            ],
        )

    return {"id": rows[0]["id"]}


async def eliminar_material_bom(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    bom_id = _int_or_none(payload.get("id"), "id")
    if not bom_id:
        raise AppError("ID del BOM es obligatorio.", 400, "VALIDATION")
    async with db.begin():
        await execute(db, "DELETE FROM bom_items WHERE id = $1", [bom_id])
    return {"ok": True}


async def _table_exists(db: AsyncSession, table_name: str) -> bool:
    row = await fetch_one(db, "SELECT to_regclass($1) AS table_ref", [table_name])
    return bool(row and row.get("table_ref"))


async def ensure_subitems_schema(db: AsyncSession) -> None:
    await execute(
        db,
        """
        CREATE TABLE IF NOT EXISTS subitems_mueble (
            id BIGSERIAL PRIMARY KEY,
            item_id INTEGER NOT NULL REFERENCES items_proyecto(id) ON DELETE CASCADE,
            nombre TEXT NOT NULL,
            activo BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
        """,
    )
    await execute(
        db,
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_subitems_mueble_item_nombre_activo
        ON subitems_mueble (item_id, LOWER(TRIM(nombre)))
        WHERE activo = TRUE
        """,
    )

    if await _table_exists(db, "piezas_mueble"):
        await execute(
            db,
            """
            ALTER TABLE piezas_mueble
            ADD COLUMN IF NOT EXISTS sub_item_id BIGINT REFERENCES subitems_mueble(id) ON DELETE SET NULL
            """,
        )
        await execute(
            db,
            """
            CREATE INDEX IF NOT EXISTS idx_piezas_mueble_sub_item_id
            ON piezas_mueble (sub_item_id)
            """,
        )

    await db.commit()


async def listar_subitems_por_item(db: AsyncSession, payload: dict | None = None) -> list[dict]:
    payload = payload or {}
    item_id = _int_or_none(payload.get("item_id"), "item_id")
    if not item_id:
        return []

    await ensure_subitems_schema(db)
    return await fetch_all(
        db,
        """SELECT id, item_id, nombre, activo, created_at, updated_at
           FROM subitems_mueble
           WHERE item_id = $1
             AND activo = TRUE
           ORDER BY nombre ASC, id ASC""",
        [item_id],
    )


async def crear_subitem(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    item_id = _int_or_none(payload.get("item_id"), "item_id")
    nombre = _text_or_none(payload.get("nombre"))
    if not item_id or not nombre:
        raise AppError("Item y nombre del sub item son obligatorios.", 400, "VALIDATION")

    await ensure_subitems_schema(db)
    async with db.begin():
        item = await fetch_one(db, "SELECT id FROM items_proyecto WHERE id = $1", [item_id])
        if not item:
            raise AppError("Item no encontrado.", 404, "NOT_FOUND")
        existing = await fetch_one(
            db,
            """SELECT id, item_id, nombre, activo, created_at, updated_at
               FROM subitems_mueble
               WHERE item_id = $1
                 AND activo = TRUE
                 AND LOWER(TRIM(nombre)) = LOWER(TRIM($2))
               LIMIT 1""",
            [item_id, nombre],
        )
        if existing:
            rows = await execute(
                db,
                """UPDATE subitems_mueble
                   SET nombre = $1,
                       updated_at = NOW()
                   WHERE id = $2
                   RETURNING id, item_id, nombre, activo, created_at, updated_at""",
                [nombre, existing["id"]],
            )
        else:
            rows = await execute(
                db,
                """INSERT INTO subitems_mueble (item_id, nombre, activo, created_at, updated_at)
                   VALUES ($1, $2, TRUE, NOW(), NOW())
                   RETURNING id, item_id, nombre, activo, created_at, updated_at""",
                [item_id, nombre],
            )
    return rows[0]


async def listar_piezas_por_item(db: AsyncSession, payload: dict | None = None) -> list[dict]:
    payload = payload or {}
    item_id = _int_or_none(payload.get("item_id"), "item_id")
    sub_item_id = _int_or_none(payload.get("sub_item_id"), "sub_item_id")
    sin_sub_item = to_bool(payload.get("sin_sub_item")) if payload.get("sin_sub_item") not in (None, "") else False
    if not item_id:
        return []

    await ensure_subitems_schema(db)
    if not await _table_exists(db, "piezas_mueble"):
        return []

    params = [item_id]
    where = ["pm.items_proyecto_id = $1"]
    if sub_item_id:
        params.append(sub_item_id)
        where.append(f"pm.sub_item_id = ${len(params)}")
    elif sin_sub_item:
        where.append("pm.sub_item_id IS NULL")

    return await fetch_all(
        db,
        f"""SELECT pm.*,
                  sm.nombre AS sub_item_nombre
           FROM piezas_mueble pm
           LEFT JOIN subitems_mueble sm ON sm.id = pm.sub_item_id
           WHERE {' AND '.join(where)}
           ORDER BY sm.nombre ASC NULLS FIRST,
                    pm.pieza ASC NULLS LAST,
                    pm.id ASC""",
        params,
    )


async def asignar_subitem_pieza(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    pieza_id = _int_or_none(payload.get("pieza_mueble_id") or payload.get("id"), "pieza_mueble_id")
    item_id = _int_or_none(payload.get("item_id"), "item_id")
    sub_item_id = _int_or_none(payload.get("sub_item_id"), "sub_item_id")
    if not pieza_id or not item_id:
        raise AppError("Pieza e item son obligatorios.", 400, "VALIDATION")

    await ensure_subitems_schema(db)
    if not await _table_exists(db, "piezas_mueble"):
        raise AppError("La tabla piezas_mueble no existe.", 404, "PIECES_TABLE_NOT_FOUND")

    async with db.begin():
        if sub_item_id:
            subitem = await fetch_one(
                db,
                """SELECT id
                   FROM subitems_mueble
                   WHERE id = $1
                     AND item_id = $2
                     AND activo = TRUE""",
                [sub_item_id, item_id],
            )
            if not subitem:
                raise AppError("Sub item no encontrado para este item.", 404, "SUBITEM_NOT_FOUND")

        rows = await execute(
            db,
            """UPDATE piezas_mueble
               SET sub_item_id = $1
               WHERE id = $2
                 AND items_proyecto_id = $3
               RETURNING id""",
            [sub_item_id, pieza_id, item_id],
        )

    if not rows:
        raise AppError("Pieza no encontrada para este item.", 404, "PIECE_NOT_FOUND")
    return {"ok": True, "id": rows[0]["id"]}


async def listar_materiales(db: AsyncSession, filters: dict | None = None) -> list[dict]:
    filters = filters or {}

    params = []
    clauses = []

    if filters.get("categoria"):
        params.append(filters["categoria"])
        clauses.append(f"mc.categoria = ${len(params)}")

    if filters.get("activo") not in (None, ""):
        params.append(to_bool(filters["activo"]))
        clauses.append(f"mc.activo = ${len(params)}")

    if filters.get("q"):
        params.append(f"%{filters['q']}%")
        q_param = f"${len(params)}"
        clauses.append(
            "(" +
            f"mc.nombre ILIKE {q_param} OR mt.material ILIKE {q_param} OR mt.sustrato ILIKE {q_param} OR "
            f"mt.formato ILIKE {q_param} OR mt.calibre ILIKE {q_param} OR mt.proveedor ILIKE {q_param} OR "
            f"mca.sustrato ILIKE {q_param} OR mca.calibre ILIKE {q_param} OR mca.proveedor ILIKE {q_param} OR "
            f"mh.referencia ILIKE {q_param} OR mh.proveedor ILIKE {q_param}" +
            ")"
        )

    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""

    return await fetch_all(
        db,
        f"""SELECT mc.id, mc.nombre, mc.categoria, mc.unidad_medida,
                  mc.descripcion, mc.activo,
                  mt.material, mt.sustrato AS tablero_sustrato,
                  mt.formato AS tablero_formato, mt.calibre AS tablero_calibre,
                  mt.costo AS tablero_costo, mt.proveedor AS tablero_proveedor,
                  mca.sustrato AS canto_sustrato, mca.proveedor AS canto_proveedor,
                  mca.calibre AS canto_calibre, mca.costo AS canto_costo,
                  mh.referencia AS herraje_referencia, mh.proveedor AS herraje_proveedor,
                  mh.costo AS herraje_costo,
                  COUNT(bi.id)::INTEGER AS referencias_bom
           FROM materiales_catalogo mc
           LEFT JOIN materiales_tableros mt ON mt.material_id = mc.id
           LEFT JOIN materiales_cantos mca ON mca.material_id = mc.id
           LEFT JOIN materiales_herrajes mh ON mh.material_id = mc.id
           LEFT JOIN bom_items bi ON bi.material_id = mc.id
           {where_sql}
           GROUP BY mc.id, mt.material, mt.sustrato, mt.formato, mt.calibre, mt.costo, mt.proveedor,
                    mca.sustrato, mca.proveedor, mca.calibre, mca.costo,
                    mh.referencia, mh.proveedor, mh.costo
           ORDER BY mc.categoria ASC, mc.nombre ASC, mc.id ASC""",
        params,
    )


async def guardar_material(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    data = {
        "id": _int_or_none(payload.get("id"), "id"),
        "nombre": clean(payload.get("nombre")),
        "categoria": clean(payload.get("categoria")),
        "unidad_medida": clean(payload.get("unidad_medida")),
        "descripcion": clean(payload.get("descripcion")),
        "activo": True if payload.get("activo") is None else to_bool(payload.get("activo")),
        "material": clean(payload.get("material")),
        "sustrato": clean(payload.get("sustrato")),
        "formato": clean(payload.get("formato")),
        "calibre": clean(payload.get("calibre")),
        "costo": decimal_or_none(payload.get("costo")),
        "proveedor": clean(payload.get("proveedor")),
        "referencia": clean(payload.get("referencia")),
    }

    nombre_canonico = _nombre_canonico_por_categoria(data)
    unidad = data["unidad_medida"] or _unidad_por_categoria(data["categoria"])
    if data["categoria"] == "canto" and unidad and unidad.lower() == "ml":
        unidad = "m"

    if not nombre_canonico or not data["categoria"] or not unidad:
        raise AppError(
            "Nombre/material base, categoría y unidad son obligatorios.",
            400,
            "VALIDATION",
        )

    async with db.begin():
        material_id = data["id"]
        creado = False

        if material_id:
            rows = await execute(
                db,
                """UPDATE materiales_catalogo
                   SET nombre = $1,
                       categoria = $2,
                       unidad_medida = $3,
                       descripcion = $4,
                       activo = $5
                   WHERE id = $6
                   RETURNING id""",
                [
                    nombre_canonico,
                    data["categoria"],
                    unidad,
                    data["descripcion"],
                    data["activo"],
                    material_id,
                ],
            )
            if not rows:
                raise AppError("Material no encontrado.", 404, "NOT_FOUND")
            material_id = rows[0]["id"]
        else:
            material_id = await _find_material_id_by_variant(db, data)

        if material_id and not data["id"]:
            rows = await execute(
                db,
                """UPDATE materiales_catalogo
                   SET nombre = $1,
                       categoria = $2,
                       unidad_medida = $3,
                       descripcion = $4,
                       activo = $5
                   WHERE id = $6
                   RETURNING id""",
                [
                    nombre_canonico,
                    data["categoria"],
                    unidad,
                    data["descripcion"],
                    data["activo"],
                    material_id,
                ],
            )
            if not rows:
                raise AppError("Material no encontrado.", 404, "NOT_FOUND")
        elif not material_id:
            rows = await execute(
                db,
                """INSERT INTO materiales_catalogo (nombre, categoria, unidad_medida, descripcion, activo)
                   VALUES ($1,$2,$3,$4,$5)
                   RETURNING id""",
                [nombre_canonico, data["categoria"], unidad, data["descripcion"], data["activo"]],
            )
            material_id = rows[0]["id"]
            creado = True

        await execute(db, "DELETE FROM materiales_tableros WHERE material_id = $1", [material_id])
        await execute(db, "DELETE FROM materiales_cantos WHERE material_id = $1", [material_id])
        await execute(db, "DELETE FROM materiales_herrajes WHERE material_id = $1", [material_id])

        if data["categoria"] == "tablero":
            await execute(
                db,
                """INSERT INTO materiales_tableros (
                    material_id, material, sustrato, formato, calibre, costo, proveedor
                ) VALUES ($1,$2,$3,$4,$5,$6,$7)""",
                [
                    material_id,
                    data["material"] or nombre_canonico,
                    data["sustrato"],
                    data["formato"],
                    data["calibre"],
                    data["costo"],
                    data["proveedor"],
                ],
            )

        if data["categoria"] == "canto":
            await execute(
                db,
                """INSERT INTO materiales_cantos (
                    material_id, sustrato, proveedor, calibre, costo
                ) VALUES ($1,$2,$3,$4,$5)""",
                [
                    material_id,
                    data["sustrato"] or data["nombre"] or nombre_canonico,
                    data["proveedor"],
                    data["calibre"],
                    data["costo"],
                ],
            )

        if data["categoria"] == "herraje":
            await execute(
                db,
                """INSERT INTO materiales_herrajes (
                    material_id, referencia, proveedor, costo
                ) VALUES ($1,$2,$3,$4)""",
                [
                    material_id,
                    data["referencia"] or data["nombre"] or nombre_canonico,
                    data["proveedor"],
                    data["costo"],
                ],
            )

    return {"id": material_id, "creado": creado}


async def _find_material_id_fuzzy(db: AsyncSession, material_data: dict) -> int | None:
    target_names = {
        _normalize_key(material_data.get("nombre")),
        _normalize_key(_nombre_canonico_por_categoria(material_data)),
        _normalize_key(material_data.get("material")),
        _normalize_key(material_data.get("sustrato")),
        _normalize_key(material_data.get("referencia")),
    }
    target_names.discard("")
    if not target_names:
        return None

    rows = await fetch_all(
        db,
        """SELECT mc.id, mc.nombre, mc.categoria,
                  mt.material, mt.sustrato AS tablero_sustrato, mt.calibre AS tablero_calibre, mt.proveedor AS tablero_proveedor,
                  mca.sustrato AS canto_sustrato, mca.calibre AS canto_calibre, mca.proveedor AS canto_proveedor,
                  mh.referencia AS herraje_referencia, mh.proveedor AS herraje_proveedor
           FROM materiales_catalogo mc
           LEFT JOIN materiales_tableros mt ON mt.material_id = mc.id
           LEFT JOIN materiales_cantos mca ON mca.material_id = mc.id
           LEFT JOIN materiales_herrajes mh ON mh.material_id = mc.id
           WHERE mc.categoria = $1
             AND mc.activo = TRUE
           ORDER BY mc.id ASC""",
        [material_data.get("categoria")],
    )
    for row in rows:
        candidate_values = [
            row.get("nombre"),
            row.get("material"),
            row.get("tablero_sustrato"),
            row.get("tablero_calibre"),
            row.get("tablero_proveedor"),
            row.get("canto_sustrato"),
            row.get("canto_calibre"),
            row.get("canto_proveedor"),
            row.get("herraje_referencia"),
            row.get("herraje_proveedor"),
        ]
        candidate_joined = _normalize_key(" ".join(str(value or "") for value in candidate_values))
        candidate_parts = {_normalize_key(value) for value in candidate_values if value}
        if target_names.intersection(candidate_parts) or any(
            target and (target in candidate_joined or candidate_joined in target)
            for target in target_names
        ):
            return row["id"]
    return None


async def _get_or_create_quote_material(db: AsyncSession, tipo: str | None, descripcion: str | None) -> tuple[int, bool]:
    material_data = _parse_material_from_quote(tipo, descripcion)
    material_id = await _find_material_id_by_variant(db, material_data)
    if not material_id:
        material_id = await _find_material_id_fuzzy(db, material_data)
    if material_id:
        return material_id, False

    nombre_canonico = _nombre_canonico_por_categoria(material_data)
    unidad = material_data.get("unidad_medida") or _unidad_por_categoria(material_data.get("categoria"))
    if not nombre_canonico or not material_data.get("categoria") or not unidad:
        raise AppError("No fue posible homologar o crear un material de la cotizacion.", 400, "MATERIAL_VALIDATION")

    rows = await execute(
        db,
        """INSERT INTO materiales_catalogo (nombre, categoria, unidad_medida, descripcion, activo)
           VALUES ($1,$2,$3,$4,TRUE)
           RETURNING id""",
        [nombre_canonico, material_data["categoria"], unidad, material_data.get("nombre")],
    )
    material_id = rows[0]["id"]

    if material_data["categoria"] == "tablero":
        await execute(
            db,
            """INSERT INTO materiales_tableros (
                material_id, material, sustrato, formato, calibre, costo, proveedor
            ) VALUES ($1,$2,$3,$4,$5,$6,$7)""",
            [
                material_id,
                material_data.get("material") or nombre_canonico,
                material_data.get("sustrato"),
                material_data.get("formato"),
                material_data.get("calibre"),
                material_data.get("costo"),
                material_data.get("proveedor"),
            ],
        )
    elif material_data["categoria"] == "canto":
        await execute(
            db,
            """INSERT INTO materiales_cantos (
                material_id, sustrato, proveedor, calibre, costo
            ) VALUES ($1,$2,$3,$4,$5)""",
            [
                material_id,
                material_data.get("sustrato") or material_data.get("nombre") or nombre_canonico,
                material_data.get("proveedor"),
                material_data.get("calibre"),
                material_data.get("costo"),
            ],
        )
    else:
        await execute(
            db,
            """INSERT INTO materiales_herrajes (
                material_id, referencia, proveedor, costo
            ) VALUES ($1,$2,$3,$4)""",
            [
                material_id,
                material_data.get("referencia") or material_data.get("nombre") or nombre_canonico,
                material_data.get("proveedor"),
                material_data.get("costo"),
            ],
        )

    return int(material_id), True


def _header_key(value) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9#]+", "", text)


def _to_int_count(value) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return 0


def _to_decimal_quantity(value):
    return decimal_or_none(value)


def _extract_quote_material_rows(file_bytes: bytes) -> list[dict]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise AppError("No fue posible leer el Excel de la cotizacion aprobada.", 400, "INVALID_QUOTE_FILE") from exc

    required = {"#", "producto", "material", "descripcion", "cantidad"}
    tipologia_keys = {"tipologia", "grupo"}

    for sheet in workbook.worksheets:
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True), start=1):
            headers = [_header_key(value) for value in row]
            header_set = set(headers)
            if not required.issubset(header_set) or not header_set.intersection(tipologia_keys):
                continue

            index = {header: idx for idx, header in enumerate(headers) if header}
            tipologia_idx = index.get("tipologia", index.get("grupo"))
            rows: list[dict] = []
            for data_row in sheet.iter_rows(min_row=row_number + 1, values_only=True):
                producto = _text_or_none(data_row[index["producto"]] if index["producto"] < len(data_row) else None)
                tipologia = _text_or_none(data_row[tipologia_idx] if tipologia_idx is not None and tipologia_idx < len(data_row) else None)
                material_tipo = _text_or_none(data_row[index["material"]] if index["material"] < len(data_row) else None)
                descripcion = _text_or_none(data_row[index["descripcion"]] if index["descripcion"] < len(data_row) else None)
                cantidad = _to_decimal_quantity(data_row[index["cantidad"]] if index["cantidad"] < len(data_row) else None)
                item_count = _to_int_count(data_row[index["#"]] if index["#"] < len(data_row) else None)
                if not producto and not tipologia and not material_tipo and not descripcion:
                    continue
                if not producto or not tipologia or not material_tipo or not descripcion or not cantidad or cantidad <= 0 or item_count <= 0:
                    continue
                rows.append(
                    {
                        "item_count": item_count,
                        "producto": producto,
                        "tipologia": tipologia,
                        "material_tipo": material_tipo,
                        "descripcion": descripcion,
                        "cantidad": cantidad,
                    }
                )
            if rows:
                return rows

    raise AppError(
        "No es posible crear items automaticamente: no se encontro una hoja con columnas #, Producto, Tipologia/Grupo, Material, Descripcion y Cantidad.",
        400,
        "QUOTE_MATERIALS_SHEET_NOT_FOUND",
    )


async def crear_items_desde_cotizacion_aprobada(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    proyecto_id = _int_or_none(payload.get("proyecto_id"), "proyecto_id")
    if not proyecto_id:
        raise AppError("Debes seleccionar un proyecto.", 400, "VALIDATION")

    from src.services.carpentry import documentos

    await documentos.ensure_document_table(db)
    approved_quote = await fetch_one(
        db,
        f"""SELECT id, nombre_archivo, tipo_archivo, azure_container, azure_blob_name
           FROM {documentos.DOCUMENTS_TABLE}
           WHERE proyecto_id = $1
             AND activo = TRUE
             AND aprobado = TRUE
             AND LOWER(etapa_nombre) LIKE '%cotiz%'
           ORDER BY aprobado_at DESC NULLS LAST, updated_at DESC NULLS LAST, id DESC
           LIMIT 1""",
        [proyecto_id],
    )
    if not approved_quote:
        raise AppError("Debes aprobar una cotizacion antes de crear items automaticamente.", 400, "QUOTE_APPROVAL_REQUIRED")
    filename = str(approved_quote.get("nombre_archivo") or "").lower()
    if not filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise AppError("La cotizacion aprobada debe ser un archivo Excel para crear items automaticamente.", 400, "INVALID_QUOTE_FILE")

    document_bytes, _, _ = await documentos.descargar_documento(db, int(approved_quote["id"]))
    quote_rows = _extract_quote_material_rows(document_bytes)

    grouped: dict[tuple[str, str], dict] = {}
    for row in quote_rows:
        key = (_normalize_key(row["producto"]), _normalize_key(row["tipologia"]))
        current = grouped.setdefault(
            key,
            {
                "producto": row["producto"],
                "tipologia": row["tipologia"],
                "item_count": row["item_count"],
                "materials": [],
            },
        )
        current["item_count"] = max(current["item_count"], row["item_count"])
        current["materials"].append(row)

    if db.in_transaction():
        await db.rollback()

    created_items = 0
    created_materials = 0
    bom_rows = 0
    async with db.begin():
        for group in grouped.values():
            item_rows = await execute(
                db,
                """INSERT INTO items_proyecto (proyecto_id, nombre, tipologia)
                   SELECT $1, $2, $3
                   FROM generate_series(1, $4)
                   RETURNING id""",
                [proyecto_id, group["producto"], group["tipologia"], group["item_count"]],
            )
            item_ids = [row["id"] for row in item_rows]
            created_items += len(item_ids)

            material_totals: dict[int, dict] = {}
            for material_row in group["materials"]:
                material_id, material_created = await _get_or_create_quote_material(
                    db,
                    material_row["material_tipo"],
                    material_row["descripcion"],
                )
                if material_created:
                    created_materials += 1
                material_totals.setdefault(
                    material_id,
                    {
                        "cantidad": 0,
                        "nota": f"{material_row['material_tipo']} - {material_row['descripcion']}",
                    },
                )
                material_totals[material_id]["cantidad"] += material_row["cantidad"]

            insert_item_ids = []
            insert_material_ids = []
            insert_quantities = []
            insert_notes = []
            for item_id in item_ids:
                for material_id, material_info in material_totals.items():
                    insert_item_ids.append(item_id)
                    insert_material_ids.append(material_id)
                    insert_quantities.append(material_info["cantidad"])
                    insert_notes.append(material_info["nota"])

            if insert_item_ids:
                await execute(
                    db,
                    """INSERT INTO bom_items (item_id, material_id, cantidad_requerida, notas)
                       SELECT *
                       FROM unnest(
                         $1::INT[],
                         $2::INT[],
                         $3::NUMERIC[],
                         $4::TEXT[]
                       ) AS src(item_id, material_id, cantidad_requerida, notas)
                       ON CONFLICT (item_id, material_id)
                       DO UPDATE SET cantidad_requerida = EXCLUDED.cantidad_requerida,
                                     notas = EXCLUDED.notas""",
                    [insert_item_ids, insert_material_ids, insert_quantities, insert_notes],
                )
                bom_rows += len(insert_item_ids)

    return {
        "ok": True,
        "items_creados": created_items,
        "materiales_creados": created_materials,
        "bom_registros": bom_rows,
        "grupos": len(grouped),
    }


async def desactivar_material(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    material_id = _int_or_none(payload.get("id"), "id")
    if not material_id:
        raise AppError("ID del material es obligatorio.", 400, "VALIDATION")

    async with db.begin():
        rows = await execute(
            db,
            """UPDATE materiales_catalogo
               SET activo = FALSE
               WHERE id = $1
               RETURNING id""",
            [material_id],
        )

    if not rows:
        raise AppError("Material no encontrado.", 404, "NOT_FOUND")

    return {"ok": True}


async def crear_items_bulk(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    proyecto_id = _int_or_none(payload.get("proyecto_id"), "proyecto_id")
    nombre = payload.get("nombre")
    cantidad = payload.get("cantidad")

    if not proyecto_id or not nombre:
        raise AppError("Proyecto y nombre son obligatorios.", 400, "VALIDATION")

    n = _int_or_zero(cantidad)

    if n < 1 or n > 500:
        raise AppError("La cantidad debe ser un número entre 1 y 500.", 400, "VALIDATION")

    async with db.begin():
        rows = await execute(
            db,
            """INSERT INTO items_proyecto (proyecto_id, nombre)
               SELECT $1, $2
               FROM generate_series(1, $3)
               RETURNING id""",
            [proyecto_id, nombre, n],
        )

    return {"creados": len(rows), "ids": [row["id"] for row in rows]}
