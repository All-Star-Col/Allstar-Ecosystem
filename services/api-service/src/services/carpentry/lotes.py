import re
import unicodedata
from io import BytesIO
from datetime import date, datetime
from typing import Any

from fastapi.concurrency import run_in_threadpool
from sqlalchemy.ext.asyncio import AsyncSession

from src.core.logging_config import get_logger
from src.services.carpentry.common import AppError, build_where, clean, execute, fetch_all, fetch_one

logger = get_logger(__name__)


def _normalize_text(value: str | None) -> str:
    text = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def _text_or_none(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def _number_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _split_apartamentos(value: Any) -> list[str]:
    text = _text_or_none(value)
    if not text:
        return []
    parts = re.split(r"\s*[-,/;]\s*|\s+y\s+", text, flags=re.IGNORECASE)
    return [part.strip() for part in parts if part and part.strip()]


def _header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9#]+", "", _normalize_text(value))


def _piece_label(value: Any) -> str:
    return _text_or_none(value) or "Pieza"


async def _ensure_lote_documento_origen_schema(db: AsyncSession) -> None:
    from src.services.carpentry import documentos

    await documentos.ensure_document_table(db)
    await execute(
        db,
        f"""
        ALTER TABLE lotes
        ADD COLUMN IF NOT EXISTS documento_orden_corte_id BIGINT REFERENCES {documentos.DOCUMENTS_TABLE}(id) ON DELETE SET NULL
        """,
    )
    await execute(
        db,
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_lotes_documento_orden_corte_id
        ON lotes (documento_orden_corte_id)
        WHERE documento_orden_corte_id IS NOT NULL
        """,
    )
    await db.commit()


async def _ensure_material_requerido_schema(db: AsyncSession) -> None:
    await execute(
        db,
        """
        CREATE TABLE IF NOT EXISTS material_requerido_lote (
            id BIGSERIAL PRIMARY KEY,
            lote_id INTEGER NOT NULL REFERENCES lotes(id) ON DELETE CASCADE,
            material_id INTEGER REFERENCES materiales_catalogo(id) ON DELETE SET NULL,
            categoria TEXT NOT NULL CHECK (categoria IN ('tablero', 'herraje')),
            nombre TEXT NOT NULL,
            cantidad NUMERIC NOT NULL CHECK (cantidad > 0),
            unidad_medida TEXT NOT NULL,
            notas TEXT,
            activo BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
        """,
    )
    await execute(
        db,
        """
        CREATE INDEX IF NOT EXISTS idx_material_requerido_lote_lote
        ON material_requerido_lote (lote_id, activo)
        """,
    )
    await db.commit()


def _requiere_maquina(proceso_nombre: str | None) -> bool:
    name = _normalize_text(proceso_nombre)
    return any(stage in name for stage in ["seccionado", "enchape", "mecanizado", "armado"])


def _es_seccionado(proceso_nombre: str | None) -> bool:
    return "seccionado" in _normalize_text(proceso_nombre)


def _int_or_none(value, field_name: str) -> int | None:
    value = clean(value)
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise AppError(f"Valor inválido para {field_name}.", 400, "VALIDATION") from exc


def _date_or_none(value, field_name: str) -> date | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text[:10]).date()
    except ValueError as exc:
        raise AppError(f"Fecha inválida en {field_name}.", 400, "VALIDATION") from exc


def _normalize_input(data: dict) -> dict:
    prioridad_raw = clean(data.get("prioridad"))
    try:
        prioridad = int(prioridad_raw or 2)
    except (TypeError, ValueError) as exc:
        raise AppError("Valor inválido para prioridad.", 400, "VALIDATION") from exc

    return {
        "proyecto_id": _int_or_none(data.get("proyecto_id"), "proyecto_id"),
        "nombre": clean(data.get("nombre")),
        "descripcion": clean(data.get("descripcion")),
        "material_ref": clean(data.get("material_ref")),
        "estado": clean(data.get("estado")) or "pendiente",
        "proceso_actual_id": _int_or_none(data.get("proceso_actual_id"), "proceso_actual_id"),
        "fecha_inicio_prog": _date_or_none(data.get("fecha_inicio_prog"), "fecha_inicio_prog"),
        "fecha_entrega_prog": _date_or_none(data.get("fecha_entrega_prog"), "fecha_entrega_prog"),
        "prioridad": prioridad,
        "notas": clean(data.get("notas")),
    }


async def listar_lotes(db: AsyncSession, filters: dict | None = None) -> list[dict]:
    filters = filters or {}
    proyecto_id = _int_or_none(filters.get("proyecto_id"), "proyecto_id")
    proceso_actual_id = _int_or_none(filters.get("proceso_actual_id"), "proceso_actual_id")
    where_sql, values, _ = build_where(
        {
            "estado": filters.get("estado"),
            "proyecto_id": proyecto_id,
            "proceso_actual_id": proceso_actual_id,
        },
        {
            "estado": "l.estado",
            "proyecto_id": "l.proyecto_id",
            "proceso_actual_id": "l.proceso_actual_id",
        },
    )

    return await fetch_all(
        db,
        f"""SELECT l.id, l.nombre, l.proyecto_id, p.nombre AS proyecto,
                  l.material_ref, l.estado, l.proceso_actual_id,
                  pr.nombre AS proceso_actual, l.prioridad,
                  l.fecha_entrega_prog,
                  (l.fecha_entrega_prog - CURRENT_DATE) AS dias_restantes
           FROM lotes l
           JOIN proyectos p ON p.id = l.proyecto_id
           LEFT JOIN procesos pr ON pr.id = l.proceso_actual_id
           {where_sql}
           ORDER BY l.prioridad ASC, l.fecha_entrega_prog ASC NULLS LAST, l.id DESC""",
        values,
    )


async def guardar_lote(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    data = _normalize_input(payload)
    lote_id = _int_or_none(payload.get("id"), "id")

    if not data["proyecto_id"] or not data["nombre"]:
        raise AppError("Proyecto y nombre del lote son obligatorios.", 400, "VALIDATION")

    async with db.begin():
        if data["proceso_actual_id"]:
            proceso_valido = await fetch_one(
                db,
                """SELECT id
                   FROM procesos
                   WHERE id = $1
                     AND orden_linea IS NOT NULL
                   LIMIT 1""",
                [data["proceso_actual_id"]],
            )
            if not proceso_valido:
                raise AppError(
                    "El proceso actual debe pertenecer a la línea de ensamblaje.",
                    400,
                    "VALIDATION",
                )

        if lote_id:
            rows = await execute(
                db,
                """UPDATE lotes
                   SET proyecto_id = $1,
                       nombre = $2,
                       descripcion = $3,
                       material_ref = $4,
                       estado = $5,
                       proceso_actual_id = $6,
                       fecha_inicio_prog = $7,
                       fecha_entrega_prog = $8,
                       prioridad = $9,
                       notas = $10
                   WHERE id = $11
                   RETURNING id""",
                [
                    data["proyecto_id"],
                    data["nombre"],
                    data["descripcion"],
                    data["material_ref"],
                    data["estado"],
                    data["proceso_actual_id"],
                    data["fecha_inicio_prog"],
                    data["fecha_entrega_prog"],
                    data["prioridad"],
                    data["notas"],
                    lote_id,
                ],
            )

            if not rows:
                raise AppError("Lote no encontrado.", 404, "NOT_FOUND")

            return {"id": rows[0]["id"], "creado": False}

        proceso_actual_id = data["proceso_actual_id"]
        if not proceso_actual_id:
            first_process = await fetch_one(
                db,
                """SELECT id
                   FROM procesos
                   WHERE orden_linea IS NOT NULL
                   ORDER BY orden_linea ASC, id ASC
                   LIMIT 1""",
            )
            proceso_actual_id = first_process["id"] if first_process else None

        insert_lote = await execute(
            db,
            """INSERT INTO lotes (
                 proyecto_id, nombre, descripcion, material_ref, estado,
                 proceso_actual_id, fecha_inicio_prog, fecha_entrega_prog,
                 prioridad, notas
               ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
               RETURNING id""",
            [
                data["proyecto_id"],
                data["nombre"],
                data["descripcion"],
                data["material_ref"],
                data["estado"],
                proceso_actual_id,
                data["fecha_inicio_prog"],
                data["fecha_entrega_prog"],
                data["prioridad"],
                data["notas"],
            ],
        )

        lote_id = insert_lote[0]["id"]

        await execute(
            db,
            """INSERT INTO lote_procesos (lote_id, proceso_id, orden, estado)
               SELECT $1,
                      p.id,
                      p.orden_linea::SMALLINT,
                      'pendiente'
               FROM procesos p
               WHERE p.orden_linea IS NOT NULL
               ORDER BY p.orden_linea ASC, p.id ASC""",
            [lote_id],
        )

    return {"id": lote_id, "creado": True}


async def obtener_lote_detalle(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    lote_id = _int_or_none(payload.get("id"), "id")
    if not lote_id:
        raise AppError("Debes indicar el lote.", 400, "VALIDATION")

    lote = await fetch_one(
        db,
        """SELECT l.id, l.proyecto_id, p.nombre AS proyecto,
                  l.nombre, l.descripcion, l.material_ref, l.estado,
                  l.proceso_actual_id, pr.nombre AS proceso_actual,
                  l.fecha_inicio_prog, l.fecha_inicio_real,
                  l.fecha_entrega_prog, l.fecha_entrega_real,
                  l.prioridad, l.notas
           FROM lotes l
           JOIN proyectos p ON p.id = l.proyecto_id
           LEFT JOIN procesos pr ON pr.id = l.proceso_actual_id
           WHERE l.id = $1""",
        [lote_id],
    )

    if not lote:
        raise AppError("Lote no encontrado.", 404, "NOT_FOUND")

    procesos = await fetch_all(
        db,
        """SELECT lp.id, lp.lote_id, lp.proceso_id, p.nombre AS proceso,
                  lp.orden, lp.fecha_programada, lp.fecha_inicio_real,
                  lp.fecha_fin_real, lp.responsable_id, lp.maquina_id,
                  per.nombre AS responsable, maq.nombre AS maquina,
                  lp.estado, lp.motivo_bloqueo, lp.notas
           FROM lote_procesos lp
           JOIN procesos p ON p.id = lp.proceso_id
           LEFT JOIN personas per ON per.id = lp.responsable_id
           LEFT JOIN maquinas maq ON maq.id = lp.maquina_id
           WHERE lp.lote_id = $1
           ORDER BY lp.orden ASC""",
        [lote_id],
    )

    items = await fetch_all(
        db,
        """SELECT ip.id, ip.proyecto_id, ip.lote_id,
                  ip.nombre, ip.tipologia, ip.tipologia AS piso, ip.apartamento, ip.estado
           FROM items_proyecto ip
           WHERE ip.lote_id = $1
           ORDER BY ip.nombre ASC, ip.tipologia ASC NULLS LAST, ip.id ASC""",
        [lote_id],
    )

    return {
        "lote": lote,
        "procesos": procesos,
        "items": items,
        "necesidadesMateriales": [],
        "materialCompleto": True,
    }


async def actualizar_lote_proceso(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    process_row_id = _int_or_none(payload.get("id"), "id")

    if not process_row_id:
        raise AppError("Falta el identificador del proceso del lote.", 400, "VALIDATION")

    async with db.begin():
        proceso = await fetch_one(
            db,
            """SELECT lp.lote_id, lp.proceso_id, p.nombre AS proceso_nombre
               FROM lote_procesos lp
               JOIN procesos p ON p.id = lp.proceso_id
               WHERE lp.id = $1
               LIMIT 1""",
            [process_row_id],
        )

        if not proceso:
            raise AppError("Proceso del lote no encontrado.", 404, "NOT_FOUND")

        maquina_id = _int_or_none(payload.get("maquina_id"), "maquina_id")
        responsable_id = _int_or_none(payload.get("responsable_id"), "responsable_id")

        if _requiere_maquina(proceso.get("proceso_nombre")) and not maquina_id:
            if _es_seccionado(proceso.get("proceso_nombre")):
                auto_maquina = await fetch_one(
                    db,
                    """SELECT id
                       FROM maquinas
                       WHERE proceso_id = $1
                         AND activo = TRUE
                       ORDER BY id ASC
                       LIMIT 1""",
                    [proceso["proceso_id"]],
                )
                if not auto_maquina:
                    raise AppError(
                        "No hay una máquina activa disponible para Seccionado.",
                        400,
                        "VALIDATION",
                    )
                maquina_id = auto_maquina["id"]
            else:
                raise AppError("Debes seleccionar una máquina para este proceso.", 400, "VALIDATION")

        if maquina_id:
            maquina_valida = await fetch_one(
                db,
                """SELECT id
                   FROM maquinas
                   WHERE id = $1
                     AND proceso_id = $2
                   LIMIT 1""",
                [maquina_id, proceso["proceso_id"]],
            )
            if not maquina_valida:
                raise AppError(
                    "La máquina seleccionada debe pertenecer al mismo proceso del lote.",
                    400,
                    "VALIDATION",
                )

        rows = await execute(
            db,
            """UPDATE lote_procesos
               SET estado = COALESCE($1, estado),
                   responsable_id = $2,
                   maquina_id = $3,
                   fecha_programada = $4,
                   fecha_inicio_real = $5,
                   fecha_fin_real = $6,
                   motivo_bloqueo = CASE WHEN COALESCE($1, estado) = 'bloqueado' THEN $7 ELSE NULL END,
                   notas = $8
               WHERE id = $9
               RETURNING lote_id, proceso_id, estado""",
            [
                clean(payload.get("estado")),
                responsable_id,
                maquina_id,
                _date_or_none(payload.get("fecha_programada"), "fecha_programada"),
                _date_or_none(payload.get("fecha_inicio_real"), "fecha_inicio_real"),
                _date_or_none(payload.get("fecha_fin_real"), "fecha_fin_real"),
                clean(payload.get("motivo_bloqueo")),
                clean(payload.get("notas")),
                process_row_id,
            ],
        )

        if not rows:
            raise AppError("Proceso del lote no encontrado.", 404, "NOT_FOUND")

        proc = rows[0]

        if proc["estado"] in ["en_proceso", "bloqueado"]:
            await execute(
                db,
                """UPDATE lotes
                   SET proceso_actual_id = $1,
                       fecha_inicio_real = CASE
                         WHEN $2 = 'en_proceso' THEN COALESCE(fecha_inicio_real, CURRENT_DATE)
                         ELSE fecha_inicio_real
                       END,
                       estado = CASE WHEN $2 = 'bloqueado' THEN 'bloqueado' ELSE 'en_produccion' END
                   WHERE id = $3""",
                [proc["proceso_id"], proc["estado"], proc["lote_id"]],
            )

        if proc["estado"] == "completado":
            pendientes = await fetch_one(
                db,
                """SELECT COUNT(*)::INTEGER AS total
                   FROM lote_procesos
                   WHERE lote_id = $1
                     AND estado <> 'completado'""",
                [proc["lote_id"]],
            )

            if pendientes and pendientes["total"] == 0:
                await execute(
                    db,
                    """UPDATE lotes
                       SET estado = 'empacado'
                       WHERE id = $1""",
                    [proc["lote_id"]],
                )

    return {"ok": True}


async def iniciar_lote_proceso(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    lote_id = _int_or_none(payload.get("lote_id"), "lote_id")

    if not lote_id:
        raise AppError("Debes indicar el lote.", 400, "VALIDATION")

    async with db.begin():
        lote = await fetch_one(
            db,
            """SELECT id, proceso_actual_id, fecha_inicio_real
               FROM lotes
               WHERE id = $1
               LIMIT 1""",
            [lote_id],
        )

        if not lote:
            raise AppError("Lote no encontrado.", 404, "NOT_FOUND")

        if lote["fecha_inicio_real"]:
            return {
                "ok": True,
                "already_started": True,
                "proceso_actual_id": lote["proceso_actual_id"],
            }

        if not lote["proceso_actual_id"]:
            raise AppError(
                "El lote no tiene proceso actual configurado.",
                400,
                "NO_CURRENT_PROCESS",
            )

        proceso_actual = await fetch_one(
            db,
            """SELECT id
               FROM lote_procesos
               WHERE lote_id = $1
                 AND proceso_id = $2
               LIMIT 1""",
            [lote_id, lote["proceso_actual_id"]],
        )

        if not proceso_actual:
            raise AppError(
                "No existe configuración del proceso actual para este lote.",
                400,
                "NO_CURRENT_PROCESS",
            )

        await execute(
            db,
            """UPDATE lote_procesos
               SET estado = 'en_proceso',
                   fecha_inicio_real = COALESCE(fecha_inicio_real, CURRENT_DATE),
                   motivo_bloqueo = NULL
               WHERE id = $1""",
            [proceso_actual["id"]],
        )

        await execute(
            db,
            """UPDATE lotes
               SET estado = 'en_produccion',
                   fecha_inicio_real = COALESCE(fecha_inicio_real, CURRENT_DATE)
               WHERE id = $1""",
            [lote_id],
        )

    return {
        "ok": True,
        "already_started": False,
        "proceso_actual_id": lote["proceso_actual_id"],
    }


async def avanzar_lote_proceso(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    lote_id = _int_or_none(payload.get("lote_id"), "lote_id")

    if not lote_id:
        raise AppError("Debes indicar el lote.", 400, "VALIDATION")

    async with db.begin():
        actual = await fetch_one(
            db,
            """SELECT lp.id, lp.lote_id, lp.proceso_id, lp.orden, lp.estado,
                      l.fecha_inicio_real AS lote_fecha_inicio_real
               FROM lote_procesos lp
               JOIN lotes l ON l.id = lp.lote_id
               WHERE lp.lote_id = $1
                 AND lp.proceso_id = l.proceso_actual_id
               LIMIT 1""",
            [lote_id],
        )

        if not actual:
            raise AppError(
                "No hay proceso actual configurado para este lote.",
                400,
                "NO_CURRENT_PROCESS",
            )
        if not actual["lote_fecha_inicio_real"]:
            raise AppError(
                "Debes iniciar producción antes de avanzar procesos.",
                400,
                "LOT_NOT_STARTED",
            )

        await execute(
            db,
            """UPDATE lote_procesos
               SET estado = 'completado',
                   fecha_fin_real = COALESCE(fecha_fin_real, CURRENT_DATE)
               WHERE id = $1""",
            [actual["id"]],
        )

        siguiente = await fetch_one(
            db,
            """SELECT id, proceso_id
               FROM lote_procesos
               WHERE lote_id = $1
                 AND orden > $2
                 AND estado IN ('pendiente','bloqueado','omitido')
               ORDER BY orden ASC
               LIMIT 1""",
            [lote_id, actual["orden"]],
        )

        if not siguiente:
            await execute(
                db,
                """UPDATE lotes
                   SET estado = 'empacado',
                       fecha_entrega_real = COALESCE(fecha_entrega_real, CURRENT_DATE)
                   WHERE id = $1""",
                [lote_id],
            )
            return {"completado": True}

        await execute(
            db,
            """UPDATE lote_procesos
               SET estado = 'en_proceso',
                   fecha_inicio_real = COALESCE(fecha_inicio_real, CURRENT_DATE),
                   motivo_bloqueo = NULL
               WHERE id = $1""",
            [siguiente["id"]],
        )

        await execute(
            db,
            """UPDATE lotes
               SET proceso_actual_id = $1,
                   estado = 'en_produccion'
               WHERE id = $2""",
            [siguiente["proceso_id"], lote_id],
        )

    return {"completado": False, "proceso_actual_id": siguiente["proceso_id"]}


async def listar_items_por_lote(db: AsyncSession, payload: dict | None = None) -> list[dict]:
    payload = payload or {}
    lote_id = _int_or_none(payload.get("lote_id"), "lote_id")

    if not lote_id:
        return []

    return await fetch_all(
        db,
        """SELECT ip.id, ip.proyecto_id, ip.lote_id,
                  ip.nombre, ip.tipologia, ip.tipologia AS piso, ip.apartamento, ip.estado
           FROM items_proyecto ip
           WHERE ip.lote_id = $1
           ORDER BY ip.nombre ASC, ip.tipologia ASC NULLS LAST, ip.id ASC""",
        [lote_id],
    )


async def _table_exists(db: AsyncSession, table_name: str) -> bool:
    row = await fetch_one(db, "SELECT to_regclass($1) AS table_ref", [table_name])
    return bool(row and row.get("table_ref"))


async def _table_columns(db: AsyncSession, table_name: str) -> set[str]:
    rows = await fetch_all(
        db,
        """SELECT a.attname AS column_name
           FROM pg_attribute a
           JOIN pg_class c ON c.oid = a.attrelid
           WHERE c.oid = to_regclass($1)
             AND a.attnum > 0
             AND NOT a.attisdropped""",
        [table_name],
    )
    return {row["column_name"] for row in rows}


async def listar_documentos_ordenes_corte(db: AsyncSession, payload: dict | None = None) -> list[dict]:
    from src.services.carpentry import documentos

    payload = payload or {}
    proyecto_id = _int_or_none(payload.get("proyecto_id"), "proyecto_id")
    await _ensure_lote_documento_origen_schema(db)

    params: list[Any] = []
    proyecto_clause = ""
    if proyecto_id:
        params.append(proyecto_id)
        proyecto_clause = f"AND d.proyecto_id = ${len(params)}"

    return await fetch_all(
        db,
        f"""SELECT d.id, d.proyecto_id, p.nombre AS proyecto,
                  d.proyecto_etapa_id, d.etapa_id, d.etapa_nombre,
                  d.titulo, d.nombre_archivo, d.tipo_archivo,
                  d.tamano_bytes, d.created_at, d.subido_por
           FROM {documentos.DOCUMENTS_TABLE} d
           JOIN proyectos p ON p.id = d.proyecto_id
           WHERE d.activo = TRUE
             {proyecto_clause}
             AND (
               LOWER(COALESCE(d.etapa_nombre, '')) LIKE '%corte%'
               OR LOWER(COALESCE(d.titulo, '')) LIKE '%corte%'
               OR LOWER(COALESCE(d.nombre_archivo, '')) LIKE '%corte%'
             )
             AND LOWER(COALESCE(d.nombre_archivo, '')) ~ '\\.(xlsx|xlsm|xltx|xltm)$'
             AND NOT EXISTS (
               SELECT 1
               FROM lotes l
               WHERE l.documento_orden_corte_id = d.id
             )
           ORDER BY d.created_at DESC, d.id DESC""",
        params,
    )


def _read_cut_order_rows_sync(file_bytes: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        header_row_number = None
        header_map: dict[str, int] = {}

        aliases = {
            "apartamento": {"apartamento", "apartamentos", "apartamentos"},
            "item": {"item", "mueble", "producto"},
            "descripcion": {"descripcion", "descrpcion", "descripcin", "descrpcin", "descrpccion"},
            "cantidad": {"cantidad", "#", "cant"},
            "largo_mm": {"largo", "largomm"},
            "ancho_mm": {"ancho", "anchomm"},
            "espesor_mm": {"espesor", "espesormm"},
            "material": {"material"},
            "referencia": {"referencia", "ref"},
        }

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
            keys = [_header_key(cell) for cell in row]
            for target, candidates in aliases.items():
                for idx, key in enumerate(keys):
                    if key in candidates or any(key.startswith(candidate) for candidate in candidates if len(candidate) > 3):
                        header_map.setdefault(target, idx)
            if {"item", "cantidad", "descripcion"}.issubset(header_map):
                header_row_number = row_idx
                break

        if not header_row_number:
            raise AppError(
                "No fue posible identificar la estructura del archivo. Debe incluir item, descripcion y cantidad.",
                400,
                "CUT_ORDER_FORMAT",
            )

        rows: list[dict[str, Any]] = []
        for excel_row, row in enumerate(
            worksheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            data: dict[str, Any] = {"fila_excel": excel_row}
            for key, idx in header_map.items():
                data[key] = row[idx] if idx < len(row) else None

            if not any(_text_or_none(data.get(k)) for k in ("apartamento", "item", "descripcion", "material", "referencia")):
                continue
            if _header_key(data.get("descripcion")) in {"l", "c"}:
                continue

            data["apartamento"] = _text_or_none(data.get("apartamento"))
            data["item"] = _text_or_none(data.get("item"))
            data["descripcion"] = _piece_label(data.get("descripcion"))
            data["material"] = _text_or_none(data.get("material"))
            data["referencia"] = _text_or_none(data.get("referencia"))
            for numeric_key in ("cantidad", "largo_mm", "ancho_mm", "espesor_mm"):
                data[numeric_key] = _number_or_none(data.get(numeric_key))
            rows.append(data)
    finally:
        workbook.close()

    return rows


async def _project_items(db: AsyncSession, proyecto_id: int) -> list[dict]:
    return await fetch_all(
        db,
        """SELECT id, proyecto_id, lote_id, nombre, tipologia, tipologia AS piso, apartamento, estado
           FROM items_proyecto
           WHERE proyecto_id = $1
           ORDER BY apartamento ASC NULLS LAST, nombre ASC, id ASC""",
        [proyecto_id],
    )


def _match_project_item(items: list[dict], apartamento: str | None, item_name: str | None) -> dict | None:
    apt_key = _normalize_text(apartamento)
    item_key = _normalize_text(item_name)
    if not apt_key or not item_key:
        return None

    same_apartment = [item for item in items if _normalize_text(item.get("apartamento")) == apt_key]
    for item in same_apartment:
        if _normalize_text(item.get("nombre")) == item_key:
            return item
    for item in same_apartment:
        name_key = _normalize_text(item.get("nombre"))
        if item_key in name_key or name_key in item_key:
            return item
    return None


def _make_preview_rows(raw_rows: list[dict], items: list[dict]) -> list[dict]:
    preview_rows: list[dict] = []
    row_id = 1

    for raw in raw_rows:
        apartamentos = _split_apartamentos(raw.get("apartamento")) or [_text_or_none(raw.get("apartamento"))]
        apartamentos = [apt for apt in apartamentos if apt]
        if not apartamentos:
            apartamentos = [None]

        cantidad_total = raw.get("cantidad")
        per_apartment: float | None = cantidad_total
        distribution_issue = False
        if len(apartamentos) > 1 and cantidad_total is not None:
            divided = cantidad_total / len(apartamentos)
            if float(divided).is_integer():
                per_apartment = float(int(divided))
            else:
                per_apartment = None
                distribution_issue = True

        for apartamento in apartamentos:
            item = _match_project_item(items, apartamento, raw.get("item"))
            issues: list[str] = []
            if not item:
                issues.append("ITEM_NO_ENCONTRADO")
            if distribution_issue:
                issues.append("DISTRIBUIR_CANTIDAD")
            if per_apartment is None or per_apartment <= 0:
                issues.append("CANTIDAD_INVALIDA")

            preview_rows.append(
                {
                    "row_id": row_id,
                    "fila_excel": raw.get("fila_excel"),
                    "apartamento": apartamento,
                    "item_nombre": item.get("nombre") if item else raw.get("item"),
                    "item_id": item.get("id") if item else None,
                    "pieza": raw.get("descripcion"),
                    "cantidad": per_apartment,
                    "cantidad_original": cantidad_total,
                    "largo_mm": raw.get("largo_mm"),
                    "ancho_mm": raw.get("ancho_mm"),
                    "espesor_mm": raw.get("espesor_mm"),
                    "material": raw.get("material"),
                    "referencia": raw.get("referencia"),
                    "subitem_nombre": raw.get("referencia"),
                    "incluir": True,
                    "issues": issues,
                }
            )
            row_id += 1

    return preview_rows


async def previsualizar_lote_desde_orden_corte(db: AsyncSession, payload: dict | None = None) -> dict:
    from src.services.carpentry import documentos

    payload = payload or {}
    documento_id = _int_or_none(payload.get("documento_id"), "documento_id")
    if not documento_id:
        raise AppError("Debes seleccionar el documento de orden de corte.", 400, "VALIDATION")

    document = await documentos.obtener_documento(db, documento_id)
    filename = str(document.get("nombre_archivo") or "").lower()
    if not filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise AppError("Por ahora la carga automatica de orden de corte requiere un Excel .xlsx.", 400, "FILE_TYPE_NOT_ALLOWED")

    file_bytes, _, _ = await documentos.descargar_documento(db, documento_id)
    raw_rows = await run_in_threadpool(_read_cut_order_rows_sync, file_bytes)
    items = await _project_items(db, int(document["proyecto_id"]))
    rows = _make_preview_rows(raw_rows, items)

    return {
        "documento": document,
        "fecha_inicio_prog": date.today().isoformat(),
        "items": [
            {
                "id": item["id"],
                "label": f"{item.get('apartamento') or '-'} - {item.get('nombre') or ''}",
                "nombre": item.get("nombre"),
                "apartamento": item.get("apartamento"),
                "tipologia": item.get("tipologia"),
            }
            for item in items
        ],
        "rows": rows,
        "resumen": {
            "filas": len(rows),
            "con_alertas": sum(1 for row in rows if row["issues"]),
            "validas": sum(1 for row in rows if not row["issues"]),
        },
    }


async def _get_or_create_subitem(db: AsyncSession, item_id: int, nombre: str | None) -> int | None:
    nombre = _text_or_none(nombre)
    if not nombre:
        return None

    existing = await fetch_one(
        db,
        """SELECT id
           FROM subitems_mueble
           WHERE item_id = $1
             AND activo = TRUE
             AND LOWER(TRIM(nombre)) = LOWER(TRIM($2))
           LIMIT 1""",
        [item_id, nombre],
    )
    if existing:
        return existing["id"]

    rows = await execute(
        db,
        """INSERT INTO subitems_mueble (item_id, nombre, activo, created_at, updated_at)
           VALUES ($1, $2, TRUE, NOW(), NOW())
           RETURNING id""",
        [item_id, nombre],
    )
    return rows[0]["id"] if rows else None


async def _insert_piece(db: AsyncSession, values: dict[str, Any], columns: set[str]) -> int | None:
    trusted_values = {key: value for key, value in values.items() if key in columns}
    if "fecha_registro" in columns:
        trusted_values["fecha_registro"] = "__NOW__"
    if "created_at" in columns:
        trusted_values["created_at"] = "__NOW__"
    if "updated_at" in columns:
        trusted_values["updated_at"] = "__NOW__"

    if not trusted_values:
        raise AppError("La tabla piezas_mueble no tiene columnas compatibles.", 500, "PIECES_SCHEMA")

    insert_columns = []
    placeholders = []
    params: list[Any] = []
    for column, value in trusted_values.items():
        insert_columns.append(column)
        if value == "__NOW__":
            placeholders.append("NOW()")
        else:
            params.append(value)
            placeholders.append(f"${len(params)}")

    rows = await execute(
        db,
        f"""INSERT INTO piezas_mueble ({", ".join(insert_columns)})
            VALUES ({", ".join(placeholders)})
            RETURNING id""",
        params,
    )
    return rows[0]["id"] if rows else None


async def crear_lote_desde_orden_corte(db: AsyncSession, payload: dict | None = None) -> dict:
    from src.services.carpentry import bom, documentos

    payload = payload or {}
    documento_id = _int_or_none(payload.get("documento_id"), "documento_id")
    fecha_entrega_prog = _date_or_none(payload.get("fecha_entrega_prog"), "fecha_entrega_prog")
    prioridad_raw = clean(payload.get("prioridad"))
    detalle = clean(payload.get("detalle"))
    rows = payload.get("rows") or []

    if not documento_id:
        raise AppError("Debes seleccionar el documento de orden de corte.", 400, "VALIDATION")
    if not fecha_entrega_prog:
        raise AppError("La fecha de entrega programada es obligatoria.", 400, "VALIDATION")
    try:
        prioridad = int(prioridad_raw)
    except (TypeError, ValueError) as exc:
        raise AppError("La prioridad es obligatoria.", 400, "VALIDATION") from exc
    if prioridad not in (1, 2, 3):
        raise AppError("La prioridad debe ser Alta, Media o Baja.", 400, "VALIDATION")
    if not isinstance(rows, list) or not rows:
        raise AppError("No hay piezas para cargar.", 400, "VALIDATION")

    await _ensure_lote_documento_origen_schema(db)
    document = await documentos.obtener_documento(db, documento_id)
    existing_lote = await fetch_one(
        db,
        """SELECT id, nombre
           FROM lotes
           WHERE documento_orden_corte_id = $1
           LIMIT 1""",
        [documento_id],
    )
    if existing_lote:
        raise AppError(
            f"Este documento ya tiene el lote {existing_lote['nombre']} creado.",
            409,
            "CUT_ORDER_ALREADY_USED",
        )
    proyecto_id = int(document["proyecto_id"])
    item_ids = {_int_or_none(row.get("item_id"), "item_id") for row in rows if row.get("incluir", True)}
    item_ids.discard(None)
    if not item_ids:
        raise AppError("Debes relacionar al menos un item valido.", 400, "VALIDATION")

    valid_items = await fetch_all(
        db,
        f"""SELECT id
           FROM items_proyecto
           WHERE proyecto_id = $1
             AND id = ANY($2::int[])""",
        [proyecto_id, list(item_ids)],
    )
    valid_item_ids = {int(item["id"]) for item in valid_items}

    prepared_rows: list[dict[str, Any]] = []
    for row in rows:
        if row.get("incluir") is False:
            continue
        item_id = _int_or_none(row.get("item_id"), "item_id")
        cantidad = _number_or_none(row.get("cantidad"))
        pieza = _piece_label(row.get("pieza"))
        if not item_id or item_id not in valid_item_ids:
            raise AppError(f"Fila {row.get('fila_excel') or row.get('row_id')}: item invalido.", 400, "VALIDATION")
        if cantidad is None or cantidad <= 0:
            raise AppError(f"Fila {row.get('fila_excel') or row.get('row_id')}: cantidad invalida.", 400, "VALIDATION")
        prepared_rows.append(
            {
                **row,
                "item_id": item_id,
                "cantidad": cantidad,
                "pieza": pieza,
                "subitem_nombre": _text_or_none(row.get("subitem_nombre") or row.get("referencia")),
            }
        )

    if not prepared_rows:
        raise AppError("No hay filas incluidas para cargar.", 400, "VALIDATION")

    if db.in_transaction():
        await db.rollback()
    await bom.ensure_subitems_schema(db)
    if not await _table_exists(db, "piezas_mueble"):
        raise AppError("La tabla piezas_mueble no existe.", 404, "PIECES_TABLE_NOT_FOUND")
    piece_columns = await _table_columns(db, "piezas_mueble")
    if db.in_transaction():
        await db.rollback()

    async with db.begin():
        first_process = await fetch_one(
            db,
            """SELECT id
               FROM procesos
               WHERE orden_linea IS NOT NULL
               ORDER BY orden_linea ASC, id ASC
               LIMIT 1""",
        )
        proceso_actual_id = first_process["id"] if first_process else None

        inserted_lote = await execute(
            db,
            """INSERT INTO lotes (
                 proyecto_id, nombre, descripcion, material_ref, estado,
                 proceso_actual_id, fecha_inicio_prog, fecha_entrega_prog,
                 prioridad, notas, documento_orden_corte_id
               ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11)
               RETURNING id""",
            [
                proyecto_id,
                document.get("nombre_archivo") or document.get("titulo") or f"Orden corte {documento_id}",
                detalle,
                None,
                "pendiente",
                proceso_actual_id,
                date.today(),
                fecha_entrega_prog,
                prioridad,
                detalle,
                documento_id,
            ],
        )
        lote_id = inserted_lote[0]["id"]

        await execute(
            db,
            """INSERT INTO lote_procesos (lote_id, proceso_id, orden, estado)
               SELECT $1,
                      p.id,
                      p.orden_linea::SMALLINT,
                      'pendiente'
               FROM procesos p
               WHERE p.orden_linea IS NOT NULL
               ORDER BY p.orden_linea ASC, p.id ASC""",
            [lote_id],
        )

        inserted_piece_ids: list[int] = []
        for row in prepared_rows:
            subitem_id = await _get_or_create_subitem(db, row["item_id"], row.get("subitem_nombre"))
            piece_id = await _insert_piece(
                db,
                {
                    "lote_id": lote_id,
                    "items_proyecto_id": row["item_id"],
                    "pieza": row["pieza"],
                    "cantidad": row["cantidad"],
                    "llegada": False,
                    "largo_mm": _number_or_none(row.get("largo_mm")),
                    "ancho_mm": _number_or_none(row.get("ancho_mm")),
                    "espesor_mm": _number_or_none(row.get("espesor_mm")),
                    "material": _text_or_none(row.get("material")),
                    "descripcion_original": _text_or_none(row.get("pieza")),
                    "archivo_origen": document.get("nombre_archivo"),
                    "hoja_origen": "orden_corte",
                    "fila_excel": _int_or_none(row.get("fila_excel"), "fila_excel"),
                    "match_tipo": "orden_corte",
                    "match_observacion": None,
                    "notas": _text_or_none(row.get("referencia")),
                    "sub_item_id": subitem_id,
                },
                piece_columns,
            )
            if piece_id:
                inserted_piece_ids.append(piece_id)

    return {
        "ok": True,
        "lote_id": lote_id,
        "piezas_creadas": len(inserted_piece_ids),
        "subitems_procesados": sum(1 for row in prepared_rows if row.get("subitem_nombre")),
    }


async def listar_material_requerido_lote(db: AsyncSession, payload: dict | None = None) -> list[dict]:
    payload = payload or {}
    lote_id = _int_or_none(payload.get("lote_id"), "lote_id")
    if not lote_id:
        return []

    await _ensure_material_requerido_schema(db)
    return await fetch_all(
        db,
        """SELECT mr.id, mr.lote_id, mr.material_id, mr.categoria,
                  mr.nombre, mr.cantidad, mr.unidad_medida, mr.notas,
                  mr.activo, mr.created_at, mr.updated_at,
                  mc.nombre AS material_catalogo
           FROM material_requerido_lote mr
           LEFT JOIN materiales_catalogo mc ON mc.id = mr.material_id
           WHERE mr.lote_id = $1
             AND mr.activo = TRUE
           ORDER BY mr.categoria ASC, mr.nombre ASC, mr.id ASC""",
        [lote_id],
    )


async def guardar_material_requerido_lote(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    row_id = _int_or_none(payload.get("id"), "id")
    lote_id = _int_or_none(payload.get("lote_id"), "lote_id")
    material_id = _int_or_none(payload.get("material_id"), "material_id")
    categoria = _text_or_none(payload.get("categoria"))
    nombre = _text_or_none(payload.get("nombre"))
    cantidad = _number_or_none(payload.get("cantidad"))
    unidad_medida = _text_or_none(payload.get("unidad_medida"))
    notas = _text_or_none(payload.get("notas"))

    if not lote_id:
        raise AppError("Debes seleccionar el lote.", 400, "VALIDATION")
    if cantidad is None or cantidad <= 0:
        raise AppError("La cantidad debe ser mayor a cero.", 400, "VALIDATION")

    await _ensure_material_requerido_schema(db)
    if db.in_transaction():
        await db.rollback()

    async with db.begin():
        lote = await fetch_one(db, "SELECT id FROM lotes WHERE id = $1", [lote_id])
        if not lote:
            raise AppError("Lote no encontrado.", 404, "NOT_FOUND")

        if material_id:
            material = await fetch_one(
                db,
                """SELECT id, nombre, categoria, unidad_medida
                   FROM materiales_catalogo
                   WHERE id = $1
                     AND activo = TRUE""",
                [material_id],
            )
            if not material:
                raise AppError("Material no encontrado.", 404, "NOT_FOUND")
            categoria = material["categoria"]
            nombre = nombre or material["nombre"]
            unidad_medida = unidad_medida or material["unidad_medida"]

        if categoria not in ("tablero", "herraje"):
            raise AppError("Material requerido solo permite tableros y herrajes.", 400, "VALIDATION")
        if not nombre or not unidad_medida:
            raise AppError("Nombre y unidad son obligatorios.", 400, "VALIDATION")

        if row_id:
            rows = await execute(
                db,
                """UPDATE material_requerido_lote
                   SET material_id = $1,
                       categoria = $2,
                       nombre = $3,
                       cantidad = $4,
                       unidad_medida = $5,
                       notas = $6,
                       updated_at = NOW()
                   WHERE id = $7
                     AND lote_id = $8
                     AND activo = TRUE
                   RETURNING id""",
                [material_id, categoria, nombre, cantidad, unidad_medida, notas, row_id, lote_id],
            )
            if not rows:
                raise AppError("Material requerido no encontrado.", 404, "NOT_FOUND")
            return {"id": rows[0]["id"], "creado": False}

        rows = await execute(
            db,
            """INSERT INTO material_requerido_lote (
                 lote_id, material_id, categoria, nombre, cantidad, unidad_medida, notas,
                 activo, created_at, updated_at
               ) VALUES ($1,$2,$3,$4,$5,$6,$7,TRUE,NOW(),NOW())
               RETURNING id""",
            [lote_id, material_id, categoria, nombre, cantidad, unidad_medida, notas],
        )
    return {"id": rows[0]["id"], "creado": True}


async def eliminar_material_requerido_lote(db: AsyncSession, payload: dict | None = None) -> dict:
    payload = payload or {}
    row_id = _int_or_none(payload.get("id"), "id")
    if not row_id:
        raise AppError("Falta el material requerido.", 400, "VALIDATION")

    await _ensure_material_requerido_schema(db)
    if db.in_transaction():
        await db.rollback()

    async with db.begin():
        rows = await execute(
            db,
            """UPDATE material_requerido_lote
               SET activo = FALSE,
                   updated_at = NOW()
               WHERE id = $1
                 AND activo = TRUE
               RETURNING id""",
            [row_id],
        )
    if not rows:
        raise AppError("Material requerido no encontrado.", 404, "NOT_FOUND")
    return {"ok": True}
